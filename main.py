from time import sleep, time
from sys import exit
from pyinputplus import inputInt
from openpyxl import load_workbook, Workbook
from os import path

hourly_rate = inputInt(prompt="Rate: ")
project_name = input("Project: ")

price = 0
task_time = 0
task_name = ''
def count(current_time,):
    global price, task_time
    try:
        while True:
            task_time = (time() - current_time)
            price = round(task_time * ((hourly_rate / 60) / 60), 2)
            print(f"${price}")
            sleep(5)
    except KeyboardInterrupt:
        print("\n")
        write_file(price)
        main()


def write_file(final_price):
    if path.exists("Work tracker.xlsx"):
        wb = load_workbook("Work tracker.xlsx")
        if project_name in wb.sheetnames:
            ws = wb[f"{project_name}"]
        else:
            ws = wb.create_sheet(f"{project_name}", 0)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = f"{project_name}"

    ws["B1"] = "=SUM(B2:B1000000)"
    row = ws.max_row + 1
    task_cell = ws.cell(row=row, column=1)
    price_cell = ws.cell(row=row, column=2)
    time_cell = ws.cell(row=row, column=3)
    rate_cell = ws.cell(row=row, column=4)
    ws[task_cell.coordinate] = f"{task_name}"
    ws[price_cell.coordinate] = final_price
    ws[time_cell.coordinate] = round(task_time / 60, 2)
    ws[rate_cell.coordinate] = f"r: ${hourly_rate}"
    with open(f"backup{project_name}.txt", 'a') as f:
        f.write(f"{task_name}: ${price}\n")
        f.close()
    wb.save('Work tracker.xlsx')


def main():
    global task_name, price
    price = 0
    task_name = input("Task: ")
    if task_name == 'q':
        exit()
    starting_time = time()
    count(starting_time)


main()
